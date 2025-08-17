import os
import json
import subprocess
import shutil
import threading
import requests
import time
from bs4 import BeautifulSoup
import pandas as pd
import json
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from parse_html import parse_html
            


def delete_folder(folder_path):
    # Check if the folder exists before attempting to delete it
    if os.path.isdir(folder_path):
        try:
            shutil.rmtree(folder_path)
            print(f"Folder '{folder_path}' and its contents deleted successfully.")
        except OSError as e:
            print(f"Error: {folder_path} : {e.strerror}")
    else:
        print(f"Folder '{folder_path}' does not exist.")




def update_xlsx(tweet):
    csv_file_path = 'tweets.csv'
    excel_file_path = 'tweets.xlsx'
    """
    Updates the CSV with a new tweet and regenerates the formatted Excel file.
    
    tweet: dict
        Example:
        {
            "tweet_text": "Some tweet text",
            "date": "2020-09-03 08:04:00",
            "image": True,
            "quote": False,
            "reply": True,
            "mentions": "@someone",
            "username": "user123",
            "link": "https://twitter.com/user123/status/..."
        }
    """

    # Ensure CSV exists, otherwise create with headers
    file_exists = os.path.isfile(csv_file_path)
    df_new = pd.DataFrame([tweet])

    if file_exists:
        df = pd.read_csv(csv_file_path)
        df = pd.concat([df, df_new], ignore_index=True)
    else:
        df = df_new

    # Reorder columns
    desired_order = ["tweet_text", "date", "image", "quote", "reply", "mentions", "username", "link"]
    df = df[desired_order]

    # Save updated CSV
    df.to_csv(csv_file_path, index=False)

    # Save to Excel first
    df.to_excel(excel_file_path, index=False)

    # Load workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Define fills (colors)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

    # Format columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_values = df[col_name].astype(str)
        max_len = max(col_values.map(len).max(), len(col_name))

        # Special case for tweet_text
        if col_name == "tweet_text":
            desired_width = max_len // 3 + 5
            wrap = True
        else:
            desired_width = min(max_len + 5, 50)  # cap width
            wrap = False

        ws.column_dimensions[col_letter].width = desired_width

        # Apply cell formatting
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=wrap, vertical="bottom")

                # Color logic
                if col_name == "image" and str(cell.value).upper() == "TRUE":
                    cell.fill = green_fill
                elif col_name == "quote" and str(cell.value).upper() == "TRUE":
                    cell.fill = yellow_fill
                elif col_name == "reply" and str(cell.value).upper() == "TRUE":
                    cell.fill = blue_fill

    # Save workbook
    wb.save(excel_file_path)

    print(f"Tweet added and '{excel_file_path}' updated successfully with formatting.")



def download_with_wmd(url, timestamp, output_dir="archive"):
    os.makedirs(output_dir, exist_ok=True)

    # Safe filename (like earlier)
    safe_name = url.replace("https://", "").replace("http://", "").replace("/", "_")
    file_name = f"{timestamp}_{safe_name}.html"
    final_file = os.path.join(output_dir, file_name)

    # Skip if already saved
    if os.path.exists(final_file):
        print(f"Skipping (already saved): {final_file}")
        return False

    # Temporary folder for downloader
    # tmp_dir = os.path.join(output_dir, "tmp_dl")
    tmp_dir = output_dir + "/tmp_dl"
    if os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir)  # clean old runs
    os.makedirs(tmp_dir, exist_ok=True)

    # Run wayback_machine_downloader
    cmd = f'wayback_machine_downloader "{url}" -e -d "{tmp_dir}"'
    print(f"Running: {cmd}")
    
    try:
        subprocess.run(cmd, shell=True, check=True)
    except subprocess.CalledProcessError as e:
        print(f"Downloader failed for {url}: {e}")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return
    # return True
    # Find index.html inside tmp_dir
    # index_path = None
    open_temp_dir = tmp_dir + "/NekroVEVO/status"
    files = os.listdir(open_temp_dir)
    file = files[0]
    with open(f"{open_temp_dir}/{file}/index.html" , encoding="utf-8") as file:
        html = file.read()

    with open(final_file, "w", encoding="utf-8") as f:
        f.write(html)
    
    folder_path = f"{open_temp_dir}"
    try:
        # os.rmdir(folder_path)
        delete_folder(folder_path)
        print(f"Folder '{folder_path}' deleted successfully.")
    except OSError as e:
        print(f"Error deleting folder '{folder_path}': {e}")
    
    soup = BeautifulSoup(html, "html.parser")
    tweet = parse_html(soup,file_name)
    update_xlsx(tweet)
    return True




def process_json_file(json_path, output_dir="archive"):
    # Load known error URLs (skip if file not found)
    try:
        with open("error_tweets.txt", "r", encoding="utf-8") as f:
            error_tweets = f.read().splitlines()
    except FileNotFoundError:
        error_tweets = []

    # Use a set for faster lookups
    error_urls = set(error_tweets)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Skip header row
    for row in data[1:]:
        original_url = row[0]
        timestamp = row[2]

        # Skip if URL is already in errors
        if original_url in error_urls:
            print(f"Skipping {original_url}, already in error log.")
            continue

        try:
            download_with_wmd(original_url, timestamp, output_dir)
        except Exception as e:
            print(f"Error processing row: {e}")
            traceback.print_exc()
            error_message = traceback.format_exc()

            # Append only if not already in log
            if original_url not in error_urls:
                error_urls.add(original_url)
                error_tweets.append(original_url)
                error_tweets.append(error_message)

                with open("error_tweets.txt", "w", encoding="utf-8") as f:
                    f.write("\n".join(error_tweets))



# Example usage
if __name__ == "__main__":
    process_json_file("nekrovevo_captures.json", "archive")
    # files = os.listdir("archive")
    # print(f"The are {len(files)} files")
